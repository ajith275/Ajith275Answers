from openpyxl import load_workbook

wb = load_workbook('data.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

count = 0

for row in range(2, max_row + 1):
    address = ws['D' + str(row)].value  # Column D = Adrese
    city = ws['E' + str(row)].value     # Column E = Pilsēta

    if isinstance(address, str) and "Adulienas iela" in address:
        if city in ["Valmiera", "Saulkrasti"]:
            count += 1

print("Matching entries count:", count)
