from openpyxl import load_workbook

wb = load_workbook('data.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

count = 0

for row in range(2, max_row + 1):
    priority = ws['H' + str(row)].value     # Column H = Priority
    delivery_date = ws['J' + str(row)].value  # Column J = Delivery date

    if priority == "High":
        if isinstance(delivery_date, str):
            if "2015" in delivery_date:
                count += 1
        elif hasattr(delivery_date, 'year'):  # if it's a datetime object
            if delivery_date.year == 2015:
                count += 1

print("Number of High-priority 2015 deliveries:", count)
