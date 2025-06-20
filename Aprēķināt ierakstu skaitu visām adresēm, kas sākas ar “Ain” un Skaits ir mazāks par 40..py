from openpyxl import load_workbook

wb = load_workbook('data.xlsx')
ws = wb['Lapa_1']
max_row = ws.max_row

count = 0

for row in range(2, max_row + 1):
    address = ws['D' + str(row)].value
    quantity = ws['L' + str(row)].value 

    if isinstance(address, str) and address.startswith("Ain"):
        if isinstance(quantity, (int, float)) and quantity < 40:
            count += 1

print("Number of matching records:", count)
