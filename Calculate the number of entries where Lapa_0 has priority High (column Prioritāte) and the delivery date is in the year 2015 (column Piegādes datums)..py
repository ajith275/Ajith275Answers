from openpyxl import load_workbook
import math

wb = load_workbook('data.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

total_price = 0
count = 0

for row in range(2, max_row + 1):
    product = ws['I' + str(row)].value  # Column I = Produkts
    price = ws['K' + str(row)].value    # Column K = Cena

    if isinstance(product, str) and 'LaserJet' in product:
        # Try to convert price if it's a string with currency symbol
        if isinstance(price, str):
            price = price.replace('€', '').replace(',', '').strip()
            try:
                price = float(price)
            except ValueError:
                continue  # Skip if not a valid number
        if isinstance(price, (int, float)):
            total_price += price
            count += 1

if count > 0:
    average = total_price / count
    print(str(math.floor(average)))  # Round down and print as string
else:
    print("0")
