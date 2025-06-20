from openpyxl import load_workbook
import math

wb = load_workbook('data.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

total_sum = 0

for row in range(2, max_row + 1):
    client = ws['F' + str(row)].value       # Column F = Client
    quantity = ws['L' + str(row)].value     # Column L = Skaits
    total = ws['N' + str(row)].value        # Column N = Kopā (Total)

    if client == "Corporate":
        if isinstance(quantity, (int, float)) and 40 <= quantity <= 50:
            if isinstance(total, str):
                total = total.replace('€', '').replace(',', '').strip()
                try:
                    total = float(total)
                except ValueError:
                    continue
            if isinstance(total, (int, float)):
                total_sum += total

print(str(math.floor(total_sum)))
